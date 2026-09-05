#!/usr/bin/env python3
"""Reconcile the daily store against an ONS Boletim Diario workbook.

The open-data portal and the bulletin are different publications of the same
operational data, so they should agree. This script proves it (or shows where
they don't) for one bulletin file.

  python check_bulletin.py DIARIO_20-08-2026.xlsx [--out data] [--tol 0.5]

Checks
  sheets 03-07  Dados Diarios Acumulados  vs  the subsystem balance series
  sheet  09     Producao Termica por Usina vs  plant_prog / plant_verif
  sheets 23-26  Sit. Princ. Reservatorios  vs  res_level_m / res_volutil_pct
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

BAL_COLS = {1: "production_total", 2: "gen_hydro", 3: "gen_thermal",
            4: "gen_wind", 5: "gen_solar", 6: "net_interchange", 7: "load"}
SHEET_SUB = {"03": "S", "04": "SE", "05": "NE", "06": "N", "07": "SIN",
             "23": "S", "24": "SE", "25": "NE", "26": "N"}


def sheets(wb, prefix):
    return [n for n in wb.sheetnames if n.startswith(prefix)]


def read_balance(wb) -> pd.DataFrame:
    rows = []
    for pref in ("03", "04", "05", "06", "07"):
        for name in sheets(wb, pref + "-"):
            sub = SHEET_SUB[pref]
            for r in wb[name].iter_rows(min_row=9, max_col=8, values_only=True):
                if r[0] is None:
                    continue
                date = pd.to_datetime(r[0], dayfirst=True, errors="coerce")
                if pd.isna(date):
                    continue
                for i, series in BAL_COLS.items():
                    if r[i] is not None:
                        rows.append((date.normalize(), sub, "", series, float(r[i])))
    return pd.DataFrame(rows, columns=["date", "subsystem", "entity",
                                       "series", "bulletin"])


def read_plants(wb, date) -> pd.DataFrame:
    name = next((n for n in wb.sheetnames if n.startswith("09-")), None)
    if not name:
        return pd.DataFrame()
    ws, rows, started = wb[name], [], False
    for r in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        head = str(r[0] or "").strip().lower()
        if head == "usina":
            started = True
            continue
        if not started or r[0] is None:
            continue
        usina = str(r[0]).strip()
        prog, verif = r[2], r[3]
        if prog is None and verif is None:
            continue
        if prog is not None:
            rows.append((date, "", usina, "plant_prog", float(prog)))
        if verif is not None:
            rows.append((date, "", usina, "plant_verif", float(verif)))
    return pd.DataFrame(rows, columns=["date", "subsystem", "entity",
                                       "series", "bulletin"])


def read_reservoirs(wb, date) -> pd.DataFrame:
    rows = []
    for pref in ("23", "24", "25", "26"):
        for name in sheets(wb, pref + "-"):
            sub = SHEET_SUB[pref]
            for r in wb[name].iter_rows(min_row=8, max_col=9, values_only=True):
                res = str(r[1] or "").strip()
                if not res:
                    continue
                for col, series in ((3, "res_level_m"), (4, "res_volutil_pct")):
                    if isinstance(r[col], (int, float)):
                        rows.append((date, sub, res, series, float(r[col])))
    return pd.DataFrame(rows, columns=["date", "subsystem", "entity",
                                       "series", "bulletin"])


def bulletin_date(wb, path: Path) -> pd.Timestamp:
    m = re.search(r"(\d{2})[-_]?(\d{2})[-_]?(\d{4})", path.name)
    if m:
        d, mo, y = m.groups()
        return pd.Timestamp(f"{y}-{mo}-{d}")
    # fall back to the last populated date in the SIN accumulation sheet
    name = next(n for n in wb.sheetnames if n.startswith("07-"))
    last = None
    for r in wb[name].iter_rows(min_row=9, max_col=8, values_only=True):
        if r[0] is not None and r[1] is not None:
            last = pd.to_datetime(r[0], dayfirst=True, errors="coerce")
    return last


def compare(tag, bul: pd.DataFrame, store: pd.DataFrame, tol: float,
            rel: bool = True) -> bool:
    if bul.empty:
        print(f"\n[{tag}] nothing read from the workbook")
        return True
    on = ["date", "subsystem", "entity", "series"]
    if tag == "plants":                       # bulletin sheet 09 has no subsystem
        on = ["date", "entity", "series"]
        store = store.drop(columns=["subsystem"])
        bul = bul.drop(columns=["subsystem"])
    m = bul.merge(store, on=on, how="left")
    miss = m["value"].isna()

    print(f"\n[{tag}] {len(m)} bulletin values, {miss.sum()} with no match in the store")
    if miss.any():
        for _, r in m[miss].head(8).iterrows():
            who = r.get("entity") or r.get("subsystem")
            print(f"    missing: {r['series']:<18} {who}")
        if miss.sum() > 8:
            print(f"    ... and {miss.sum() - 8} more")

    m = m[~miss].copy()
    if m.empty:
        return False
    m["diff"] = (m["value"] - m["bulletin"]).abs()
    denom = m["bulletin"].abs().clip(lower=1e-6)
    m["pct"] = 100 * m["diff"] / denom
    metric = "pct" if rel else "diff"
    bad = m[m[metric] > tol]

    print(f"    matched {len(m)}  |  median {metric} {m[metric].median():.3f}"
          f"  |  p95 {m[metric].quantile(.95):.3f}  |  max {m[metric].max():.3f}")
    if len(bad):
        print(f"    {len(bad)} beyond tolerance ({tol}):")
        for _, r in bad.sort_values(metric, ascending=False).head(8).iterrows():
            who = r.get("entity") or r.get("subsystem", "")
            print(f"      {r['series']:<18} {str(who)[:24]:<24} "
                  f"bulletin {r['bulletin']:>12,.2f}   store {r['value']:>12,.2f}"
                  f"   ({r['pct']:.2f}%)")
    else:
        print("    all within tolerance")
    return len(bad) == 0 and not miss.any()


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("--out", default="data")
    ap.add_argument("--tol", type=float, default=0.5,
                    help="percent difference treated as a match (default 0.5)")
    args = ap.parse_args(argv)

    path = Path(args.workbook)
    store_path = Path(args.out) / "daily.parquet"
    if not store_path.exists():
        raise SystemExit(f"{store_path} not found - run `build` first.")
    store = pd.read_parquet(store_path)
    store["date"] = pd.to_datetime(store["date"]).dt.normalize()
    store["entity"] = store["entity"].fillna("").astype(str)
    # the store holds the four subsystems; SIN is derived the same way the
    # dashboard derives it, so sheet 07 has something to compare against
    from dashboard import add_sin
    store = add_sin(store)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    date = bulletin_date(wb, path)
    print(f"Bulletin: {path.name}   date {date.date()}")
    print(f"Store:    {store_path}  {store['date'].min().date()} .. "
          f"{store['date'].max().date()}")

    ok = True
    ok &= compare("balance sheets 03-07", read_balance(wb), store, args.tol)
    ok &= compare("plants", read_plants(wb, date), store, args.tol)
    ok &= compare("reservoirs", read_reservoirs(wb, date), store, args.tol)

    print("\nAll checks passed." if ok else
          "\nDifferences found - see above. Small residuals are normal where ONS has "
          "revised one publication and not the other; large or systematic gaps are not.")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
